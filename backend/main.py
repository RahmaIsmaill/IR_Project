from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from elasticsearch import Elasticsearch
import json, os, csv, glob, datetime, re
from typing import Optional, List

app = FastAPI(title="Mini Search Engine")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

ES_HOST     = os.getenv("ES_HOST", "http://localhost:9200")
ES_USER     = os.getenv("ES_USER", "")
ES_PASS     = os.getenv("ES_PASS", "")
ES_CLOUD_ID = os.getenv("ES_CLOUD_ID", "")

if ES_CLOUD_ID:
    es = Elasticsearch(cloud_id=ES_CLOUD_ID, basic_auth=(ES_USER, ES_PASS))
elif ES_USER:
    es = Elasticsearch(ES_HOST, basic_auth=(ES_USER, ES_PASS))
else:
    es = Elasticsearch(ES_HOST)

INDEX    = "products"
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

MAPPING = {
    "mappings": {
        "properties": {
            "name":        {"type": "text", "analyzer": "standard", "fields": {"keyword": {"type": "keyword"}}},
            "description": {"type": "text", "analyzer": "standard"},
            "content":     {"type": "text", "analyzer": "standard"},
            "category":    {"type": "keyword"},
            "brand":       {"type": "keyword"},
            "price":       {"type": "integer"},
            "file_type":   {"type": "keyword"},
            "filename":    {"type": "keyword"},
            "mod_date":    {"type": "date", "format": "yyyy-MM-dd"}
        }
    },
    "settings": {"index": {"number_of_shards": 1, "number_of_replicas": 0}}
}

# ── PARSERS ──────────────────────────────────────────────
def get_mod_date(path):
    return datetime.datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d")

def parse_json(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    items = data if isinstance(data, list) else [data]
    results = []
    for doc in items:
        text = " ".join(str(v) for v in doc.values() if isinstance(v, (str, int, float)))
        d = {k: v for k, v in doc.items()}
        d.setdefault("name", os.path.basename(path))
        d.setdefault("description", text[:300])
        d["content"]   = text
        d["file_type"] = "json"
        d["filename"]  = os.path.basename(path)
        d["mod_date"]  = get_mod_date(path)
        results.append(d)
    return results

def parse_csv(path):
    results = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        for row in reader:
            text = " ".join(str(v) for v in row.values() if v)
            d = dict(row)
            d["name"]        = list(row.values())[0] if row else os.path.basename(path)
            d["description"] = text[:300]
            d["content"]     = text
            d["file_type"]   = "csv"
            d["filename"]    = os.path.basename(path)
            d["mod_date"]    = get_mod_date(path)
            results.append(d)
    return results

def parse_txt(path):
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    return [{"name": os.path.basename(path), "description": text[:300],
             "content": text, "file_type": "txt",
             "filename": os.path.basename(path), "mod_date": get_mod_date(path)}]

def parse_pdf(path):
    try:
        import fitz
        doc  = fitz.open(path)
        text = " ".join(page.get_text() for page in doc)
    except ImportError:
        text = "[PDF: install PyMuPDF — pip install PyMuPDF]"
    except Exception as e:
        text = f"[PDF error: {e}]"
    return [{"name": os.path.basename(path), "description": text[:300],
             "content": text, "file_type": "pdf",
             "filename": os.path.basename(path), "mod_date": get_mod_date(path)}]

def parse_xlsx(path):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        results = []
        for sheet in wb.sheetnames:
            ws   = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows: continue
            headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
            for row in rows[1:]:
                d    = {headers[i]: str(v) for i, v in enumerate(row) if v is not None}
                text = " ".join(d.values())
                d["name"]        = list(d.values())[0] if d else os.path.basename(path)
                d["description"] = text[:300]
                d["content"]     = text
                d["file_type"]   = "xlsx"
                d["filename"]    = os.path.basename(path)
                d["mod_date"]    = get_mod_date(path)
                results.append(d)
        return results
    except Exception as e:
        return [{"name": os.path.basename(path), "description": str(e),
                 "content": "", "file_type": "xlsx",
                 "filename": os.path.basename(path), "mod_date": get_mod_date(path)}]

PARSERS = {"json": parse_json, "csv": parse_csv, "txt": parse_txt,
           "pdf": parse_pdf, "xlsx": parse_xlsx}

def collect_docs(formats=None):
    if formats is None:
        formats = list(PARSERS.keys())
    docs = []
    for fmt in formats:
        for path in glob.glob(os.path.join(DATA_DIR, f"*.{fmt}")):
            try:
                parsed = PARSERS[fmt](path)
                docs.extend(parsed)
                print(f"  {os.path.basename(path)} -> {len(parsed)} doc(s)")
            except Exception as e:
                print(f"  ERROR {os.path.basename(path)}: {e}")
    return docs

def do_index(formats=None):
    if es.indices.exists(index=INDEX):
        es.indices.delete(index=INDEX)
    es.indices.create(index=INDEX, body=MAPPING)
    docs = collect_docs(formats)
    for i, doc in enumerate(docs):
        if "price" in doc:
            try:    doc["price"] = int(float(str(doc["price"]).replace(",", "")))
            except: del doc["price"]
        es.index(index=INDEX, id=i, document=doc)
    es.indices.refresh(index=INDEX)
    return docs

# ── STARTUP ──────────────────────────────────────────────
@app.on_event("startup")
def load_data():
    try:
        docs = do_index()
        print(f"Indexed {len(docs)} documents total")
    except Exception as e:
        print(f"Startup error: {e}")

# ── REBUILD ──────────────────────────────────────────────
@app.post("/rebuild")
def rebuild(formats: Optional[List[str]] = None):
    try:
        selected = formats if formats else list(PARSERS.keys())
        docs     = do_index(selected)
        summary  = {}
        for d in docs:
            ft = d.get("file_type", "unknown")
            summary[ft] = summary.get(ft, 0) + 1
        return {"status": "success", "total": len(docs),
                "formats": selected, "by_type": summary}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── SEARCH ───────────────────────────────────────────────
@app.get("/search")
def search(
    q: str,
    page: int = 1,
    size: int = 5,
    category: Optional[str] = None,
    file_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    min_price: Optional[int] = None,
    max_price: Optional[int] = None,
):
    from_idx = (page - 1) * size

    main_query = {
        "query_string": {
            "query": q,
            "fields": ["name^3", "description^2", "content", "category", "brand"],
            "default_operator": "OR",
            "fuzziness": "AUTO",
            "phrase_slop": 2,
            "allow_leading_wildcard": True,
            "analyze_wildcard": True
        }
    }

    filters = []
    if category:  filters.append({"term": {"category": category}})
    if file_type: filters.append({"term": {"file_type": file_type}})
    if min_price is not None or max_price is not None:
        pr = {}
        if min_price is not None: pr["gte"] = min_price
        if max_price is not None: pr["lte"] = max_price
        filters.append({"range": {"price": pr}})
    if date_from or date_to:
        dr = {}
        if date_from: dr["gte"] = date_from
        if date_to:   dr["lte"] = date_to
        filters.append({"range": {"mod_date": dr}})

    full_query = {"bool": {"must": [main_query], "filter": filters}} if filters else main_query

    highlight = {
        "fields": {
            "name":        {"number_of_fragments": 0},
            "description": {"number_of_fragments": 1, "fragment_size": 150},
            "content":     {"number_of_fragments": 1, "fragment_size": 150}
        },
        "pre_tags": ["<mark>"], "post_tags": ["</mark>"]
    }

    try:
        res   = es.search(index=INDEX, body={"query": full_query, "highlight": highlight,
                                              "from": from_idx, "size": size, "track_total_hits": True})
        total = res["hits"]["total"]["value"]
        hits  = res["hits"]["hits"]

        results = []
        for hit in hits:
            src = hit["_source"]
            hlt = hit.get("highlight", {})
            hn  = hlt.get("name", [src.get("name", "")])[0]
            hd  = hlt.get("description", hlt.get("content",
                  [src.get("description", src.get("content", ""))]))[0]
            results.append({
                "name":           src.get("name"),
                "description":    src.get("description") or src.get("content", "")[:200],
                "category":       src.get("category"),
                "brand":          src.get("brand"),
                "price":          src.get("price"),
                "file_type":      src.get("file_type"),
                "filename":       src.get("filename"),
                "mod_date":       src.get("mod_date"),
                "score":          round(hit["_score"], 3),
                "highlight_name": hn,
                "highlight_desc": hd
            })

        suggestion = None
        if total == 0:
            try:
                t_res = es.search(index=INDEX, body={"suggest": {"text": q,
                    "term_suggest": {"term": {"field": "name", "suggest_mode": "popular"}}}})
                parts = []
                for token in t_res.get("suggest", {}).get("term_suggest", []):
                    opts = token.get("options", [])
                    parts.append(opts[0]["text"] if opts else token["text"])
                candidate = " ".join(parts)
                if candidate.lower() != q.lower():
                    suggestion = candidate
            except:
                pass

        return {"total": total, "page": page,
                "total_pages": (total + size - 1) // size,
                "results": results, "suggestion": suggestion}
    except Exception as e:
        return {"error": str(e), "total": 0, "results": [], "page": 1, "total_pages": 0}

# ── STATS ────────────────────────────────────────────────
@app.get("/stats")
def stats():
    try:
        total = es.count(index=INDEX)["count"]
        agg   = es.search(index=INDEX, body={"size": 0, "aggs": {
            "by_type":     {"terms": {"field": "file_type", "size": 20}},
            "by_category": {"terms": {"field": "category",  "size": 20}}
        }})
        by_type     = {b["key"]: b["doc_count"] for b in agg["aggregations"]["by_type"]["buckets"]}
        by_category = {b["key"]: b["doc_count"] for b in agg["aggregations"]["by_category"]["buckets"]}

        # Top 10 frequent terms — real word frequency counted in Python
        top_terms = []
        try:
            hits = es.search(index=INDEX, body={
                "query": {"match_all": {}},
                "size": 500,
                "_source": ["name", "description"]
            })["hits"]["hits"]

            STOP = {
                "the","a","an","and","or","of","in","to","for","with","on","at","by",
                "from","is","are","was","were","be","been","has","have","had","it",
                "its","as","this","that","not","but","can","will","all","also","more",
                "than","their","they","which","when","who","so","if","about","into",
                "up","out","do","does","did","some","any","each","over","after","new",
                "one","two","three","per","via","inc","get","set","use","used","very",
            }

            freq = {}
            for h in hits:
                src  = h["_source"]
                text = " ".join(filter(None, [
                    str(src.get("name", "")),
                    str(src.get("description", ""))
                ]))
                for w in re.findall(r"[a-zA-Z]{3,}", text.lower()):
                    if w not in STOP and not re.match(r"^user_?\d+$", w):
                        freq[w] = freq.get(w, 0) + 1

            top_terms = [
                {"term": w, "count": c}
                for w, c in sorted(freq.items(), key=lambda x: -x[1])[:10]
            ]
        except Exception:
            pass

        return {"total_documents": total, "by_file_type": by_type,
                "by_category": by_category, "top_terms": top_terms}
    except Exception as e:
        return {"error": str(e)}

# ── INDEX INFO ───────────────────────────────────────────
@app.get("/index-info")
def index_info():
    try:
        info = es.indices.stats(index=INDEX)
        return {"status": "connected", "index": INDEX,
                "doc_count":  info["indices"][INDEX]["total"]["docs"]["count"],
                "size_bytes": info["indices"][INDEX]["total"]["store"]["size_in_bytes"]}
    except Exception as e:
        return {"status": "error", "message": str(e)}